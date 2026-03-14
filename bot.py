import datetime
from openpyxl import load_workbook, Workbook
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes, filters
import os

EXCEL_FILE = r"C:\Users\Hype\OneDrive\Documents\DIMAS PRAST\MY LIFE\HITUNG MUMET.xlsx"

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Detail Pengeluaran"
        ws.append(["Tanggal","Kategori","Deskripsi","Jumlah"])
        wb.save(EXCEL_FILE)

def save_expense(kategori, jumlah, deskripsi):
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Detail Pengeluaran"]
    tanggal = datetime.date.today().isoformat()
    ws.append([tanggal, kategori, deskripsi, int(jumlah)])
    wb.save(EXCEL_FILE)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Bot pencatat keuangan aktif.\n"
        "Contoh:\n"
        "/out makan 25000 nasi padang"
    )

async def out(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        kategori = context.args[0]
        jumlah = context.args[1]
        deskripsi = " ".join(context.args[2:]) if len(context.args) > 2 else ""
        save_expense(kategori, jumlah, deskripsi)
        await update.message.reply_text("Pengeluaran berhasil dicatat.")
    except:
        await update.message.reply_text("Format salah.")

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        parts = update.message.text.split()
        kategori = parts[0]
        jumlah = parts[1]
        deskripsi = " ".join(parts[2:]) if len(parts) > 2 else ""
        save_expense(kategori, jumlah, deskripsi)
        await update.message.reply_text("Pengeluaran tersimpan.")
    except:
        await update.message.reply_text("Format pesan tidak dikenali.")

def main():
    init_excel()

    TOKEN = "8660138743:AAF0DO5VWnJO7IHPTn9fJM1T2MDCaJ76jpA"

    app = ApplicationBuilder().token(TOKEN).build()
    
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("out", out))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))

    print("Bot running...")
    app.run_polling()

if __name__ == "__main__":
    main()

if update.message:
    text = update.message.text
